import os
from openpyxl import Workbook, load_workbook

# Function to read the last entered roll number from a file
def read_last_roll_number(file_path):
    try:
        with open(file_path, 'r') as file:
            last_roll_number = int(file.readline().strip())
            return last_roll_number
    except FileNotFoundError:
        return None

# Function to save the last entered roll number to a file
def save_last_roll_number(file_path, roll_no):
    with open(file_path, 'w') as file:
        file.write(str(roll_no))

# Function to validate and get a number of a specified length from the user
def get_valid_number(prompt, length):
    while True:
        number = input(prompt)
        if len(number) == length and number.isdigit():
            return number
        print(f"Invalid input! Please enter a {length}-digit number.")

# Function to map course selection abbreviations to full names
def map_course_selection(abbreviation):
    course_mapping = {
        'BSCN': 'BSC Nursing',
        'DP': 'DPharma'
    }
    return course_mapping.get(abbreviation.upper(), abbreviation)

# Function to map hostel requirement input to 'Yes' or 'No'
def map_hostel_requirement(input_value):
    hostel_mapping = {
        'Y': 'Yes',
        'N': 'No'
    }
    return hostel_mapping.get(input_value.upper(), input_value)

# Declare session_year as a constant
session_year = "2023-2024"

# Path to the existing or new file to store the student data
output_path = "D:/pythonexp/Personal/student_data.xlsx"  # Replace with your desired output path and filename

# Check if the output file already exists
if os.path.exists(output_path):
    # Load the existing data into a Workbook
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    last_roll = worksheet.cell(row=worksheet.max_row, column=2).value if worksheet.max_row > 1 else 0
else:
    # If the file doesn't exist, create a new Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append([
        'Session_year', 'Roll_no', 'Name', 'Father_name', 'Mother_name', 'Date_of_birth',
        'Personal_phone', 'Guardian_phone', 'Aadhaar_num', 'Bank_name', 'Bank_ifsc',
        'Bank_account_num', 'Hostel_requirement', 'Course_selection'
    ])
    last_roll = 0

data_list = []

# Collect data from the user
while True:
    last_roll += 1
    roll_no = int(input(f"Enter Roll no. (Last entered: {last_roll}): "))
    name = input("\nEnter Student name: ")
    father_name = input("\nEnter Father's name: ")
    mother_name = input("\nEnter mother's name: ")
    date_of_birth = input("Enter Date of birth: ")
    personal_phone = get_valid_number("Enter Personal phone number (10 digits): ", 10)
    guardian_phone = get_valid_number("Enter Guardian phone number (10 digits): ", 10)
    aadhaar_num = int(get_valid_number("Enter Aadhaar-card number (12 digits): ", 12))
    bank_name = input("Enter Bank name: ")
    bank_ifsc = input("Enter Bank ifsc: ")
    bank_account_num = int(input("Enter bank account no.: "))
    hostel_requirement = map_hostel_requirement(input("Require a hostel? Enter 'Y' for Yes, or 'N' for No: "))
    course_selection = map_course_selection(input("BSC Nursing (BSCN), ANM, DPharma (DP): "))
    data_list.append([session_year, roll_no, name, father_name, mother_name, date_of_birth, personal_phone,
                      guardian_phone, aadhaar_num, bank_name, bank_ifsc, bank_account_num,
                      hostel_requirement, course_selection])
    if input("Do you want to enter data for another student? (y/n): ").lower() != 'y':
        break

# Save the last entered roll number for the next run
save_last_roll_number('last_roll.txt', last_roll)

# Append the new data to the existing Worksheet
for data in data_list:
    worksheet.append(data)

# Format all columns as text to avoid scientific notation
for col in worksheet.columns:
    for cell in col:
        cell.number_format = '0'

# Save the Workbook to the Excel file
workbook.save(output_path)

print(f"Data has been exported to {output_path} successfully.")