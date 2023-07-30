import os
import subprocess
from openpyxl import Workbook, load_workbook

# Function to terminate any running Excel processes
def kill_excel_processes():
    try:
        subprocess.run(['TASKKILL', '/F', '/IM', 'EXCEL.EXE'], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error: {e}")

# Function to read the last roll number from a file
def read_last_roll_number(file_path):
    try:
        with open(file_path, 'r') as file:
            last_roll_number = int(file.readline().strip())
            return last_roll_number
    except FileNotFoundError:
        return None

# Function to save the last roll number to a file
def save_last_roll_number(file_path, roll_no):
    with open(file_path, 'w') as file:
        file.write(str(roll_no))

# Function to get a valid number with a specified length from the user
def get_valid_number(prompt, length):
    while True:
        number = input(prompt)
        if len(number) == length and number.isdigit():
            return number
        print(f"Invalid input! Please enter a {length}-digit number.")

# Function to map course abbreviation to full course name
def map_course_selection(abbreviation):
    course_mapping = {
        'BSCN': 'BSC Nursing',
        'DP': 'DPharma'
    }
    return course_mapping.get(abbreviation.upper(), abbreviation)

# Function to map hostel requirement input to 'Yes' or 'No'
def map_hostel_requirement(input_value):
    hostel_mapping = {
        'Y': 'Yes',
        'N': 'No'
    }
    return hostel_mapping.get(input_value.upper(), input_value)

output_path = "D:/pythonexp/Personal/student_data.xlsx"
session_year = "2023-2024"
data_list = []
last_roll = 0

# Check if the output file exists
if os.path.exists(output_path):
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    last_roll = worksheet.cell(row=worksheet.max_row, column=2).value if worksheet.max_row > 1 else 0
else:
    workbook = Workbook()
    worksheet = workbook.active
    # Add headers to the worksheet
    worksheet.append([
        'Session_year', 'Roll_no', 'Name', 'Father_name', 'Mother_name', 'Date_of_birth',
        'Personal_phone', 'Guardian_phone', 'Aadhaar_num', 'Bank_name', 'Bank_ifsc',
        'Bank_account_num', 'Hostel_requirement', 'Course_selection'])
if True:
    kill_excel_processes()
# Loop to get student data from the user
while True:
    last_roll += 1
    roll_no = int(input(f"Enter Roll no. (Last entered: {last_roll}): "))
    # Get student details from the user
    name = input("\nEnter Student name: ")
    father_name = input("\nEnter Father's name: ")
    mother_name = input("\nEnter mother's name: ")
    date_of_birth = "'" + input("\nEnter Date of birth: ")  # Add ' to store date_of_birth as a string
    personal_phone = get_valid_number("\nEnter Personal phone number (10 digits): ", 10)
    guardian_phone = get_valid_number("\nEnter Guardian phone number (10 digits): ", 10)
    aadhaar_num = int(get_valid_number("\nEnter Aadhaar-card number (12 digits): ", 12))
    bank_name = input("\nEnter Bank name: ")
    bank_ifsc = input("\nEnter Bank ifsc: ")
    bank_account_num = int(input("\nEnter bank account no.: "))
    hostel_requirement = map_hostel_requirement(input("\nRequire a hostel? Enter 'Y' for Yes, or 'N' for No: "))
    course_selection = map_course_selection(input("\nBSC Nursing (BSCN), ANM, DPharma (DP): "))
    data_list.append([session_year, roll_no, name, father_name, mother_name, date_of_birth, personal_phone,
                      guardian_phone, aadhaar_num, bank_name, bank_ifsc, bank_account_num,
                      hostel_requirement, course_selection])
    last_roll += 1

    if input("Do you want to enter data for another student? (y/n): ").lower() != 'y':
        break

save_last_roll_number('last_roll.txt', last_roll)

# Add data to the worksheet
for data in data_list:
    worksheet.append(data)

# Adjust column widths based on the length of the data
for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try: 
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[column].width = adjusted_width

workbook.save(output_path)
print(f"Data has been exported to {output_path} successfully.")